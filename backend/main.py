from fastapi.staticfiles import StaticFiles
from fastapi import (FastAPI, Depends,UploadFile,File)
from fastapi.responses import JSONResponse
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl import load_workbook
import random
from datetime import datetime

from backend.database import engine, get_db

from backend.models import (
Base,
Usuario,
Tema,
Pergunta,
Alternativa,
Resultado,
Resposta,
Tentativa,
Auditoria
)

from backend.schemas import (
UsuarioCreate,
TemaCreate,
PerguntaCreate,
AlternativaCreate,
RespostaCreate,
TentativaCreate,
LoginRequest
)

Base.metadata.create_all(bind=engine)

app = FastAPI()
app.mount("/frontend", StaticFiles(directory="frontend"), name="frontend")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def home():
    return FileResponse("frontend/login.html")

@app.post("/usuarios")
def criar_usuario(usuario: UsuarioCreate, db: Session = Depends(get_db)):

    novo_usuario = Usuario(
        nome=usuario.nome,
        email=usuario.email,
        senha=usuario.senha,
        avatar=usuario.avatar
    )

    db.add(novo_usuario)
    db.commit()

    return {"mensagem": "Usuário criado com sucesso"}
@app.post("/temas")
def criar_tema(tema: TemaCreate, db: Session = Depends(get_db)):

    novo_tema = Tema(
        nome=tema.nome
    )

    db.add(novo_tema)
    db.commit()

    return {"mensagem": "Tema criado com sucesso"}

@app.get("/temas")
def listar_temas(db: Session = Depends(get_db)):

    temas = db.query(Tema).all()

    return temas

@app.put("/liberar-tema/{tema_id}")
def liberar_tema(
    tema_id: int,
    db: Session = Depends(get_db)
):

    db.query(Tema).update(
        {
            "liberado": False
        }
    )

    tema = db.query(Tema).filter(
        Tema.id == tema_id
    ).first()

    if not tema:

        return {
            "mensagem":
            "Tema não encontrado."
        }

    tema.liberado = True

    novo_log = Auditoria(
        usuario_id=1,
        acao=f"Tema liberado: {tema.nome}",
        data_hora=datetime.now()
    )

    db.add(novo_log)

    db.commit()

    return {
        "mensagem":
        "Tema liberado com sucesso."
    }

@app.post("/perguntas")
def criar_pergunta(
    pergunta: PerguntaCreate,
    db: Session = Depends(get_db)
):

    nova_pergunta = Pergunta(
        enunciado=pergunta.enunciado,
        tipo=pergunta.tipo,
        tema_id=pergunta.tema_id
    )

    db.add(nova_pergunta)
    db.commit()

    return {"mensagem": "Pergunta criada com sucesso"}

@app.get("/perguntas")
def listar_perguntas(
    db: Session = Depends(get_db)
):

    perguntas = db.query(Pergunta).all()

    return perguntas

@app.post("/alternativas")
def criar_alternativa(
    alternativa: AlternativaCreate,
    db: Session = Depends(get_db)
):

    nova_alternativa = Alternativa(
        descricao=alternativa.descricao,
        correta=alternativa.correta,
        pergunta_id=alternativa.pergunta_id
    )

    db.add(nova_alternativa)
    db.commit()

    return {"mensagem": "Alternativa criada com sucesso"}

@app.get("/alternativas")
def listar_alternativas(
    db: Session = Depends(get_db)
):

    alternativas = db.query(Alternativa).all()

    return alternativas

@app.post("/iniciar-tentativa")
def iniciar_tentativa(
    tentativa: TentativaCreate,
    db: Session = Depends(get_db)
):

    quantidade_tentativas = db.query(Tentativa).filter(
        Tentativa.usuario_id == tentativa.usuario_id,
        Tentativa.concluida == True
    ).count()

    if quantidade_tentativas >= 3:

        melhor_resultado = db.query(Resultado).filter(
            Resultado.usuario_id == tentativa.usuario_id
        ).order_by(
            Resultado.percentual.desc()
        ).first()

        melhor_percentual = 0

        if melhor_resultado:
            melhor_percentual = melhor_resultado.percentual

        return {
            "bloqueado": True,
            "mensagem": "Você atingiu o limite de 3 tentativas.",
            "melhor_percentual": melhor_percentual
        }

    # Verifica se existe uma tentativa em aberto
    tentativa_aberta = db.query(Tentativa).filter(
        Tentativa.usuario_id == tentativa.usuario_id,
        Tentativa.concluida == False
    ).first()

    if tentativa_aberta:

        return {
            "bloqueado": False,
            "tentativa_id": tentativa_aberta.id
        }

    nova_tentativa = Tentativa(
        usuario_id=tentativa.usuario_id,
        concluida=False
    )

    db.add(nova_tentativa)
    db.commit()
    db.refresh(nova_tentativa)

    return {
        "bloqueado": False,
        "tentativa_id": nova_tentativa.id
    }

@app.post("/responder")
def responder(
    resposta: RespostaCreate,
    db: Session = Depends(get_db)
):

    nova_resposta = Resposta(
        usuario_id=resposta.usuario_id,
        pergunta_id=resposta.pergunta_id,
        alternativa_id=resposta.alternativa_id,
        tentativa_id=resposta.tentativa_id
    )

    db.add(nova_resposta)
    db.commit()

    alternativa = db.query(Alternativa).filter(
        Alternativa.id == resposta.alternativa_id
    ).first()

    alternativa_correta = db.query(Alternativa).filter(
        Alternativa.pergunta_id == resposta.pergunta_id,
        Alternativa.correta.is_(True)
    ).first()

    if alternativa.correta:
        return {
            "resultado": "acertou",
            "alternativa_correta": alternativa_correta.id
        }

    return {
        "resultado": "errou",
        "alternativa_correta": alternativa_correta.id
    }

@app.get("/resultado/{usuario_id}")
def resultado(usuario_id: int, db: Session = Depends(get_db)):

    ultima_tentativa = db.query(Tentativa).filter(
        Tentativa.usuario_id == usuario_id
    ).order_by(
        Tentativa.id.desc()
    ).first()

    if not ultima_tentativa:
        return {
            "mensagem": "Nenhuma tentativa encontrada."
        }

    respostas = db.query(Resposta).filter(
        Resposta.tentativa_id == ultima_tentativa.id
    ).all()

    total = len(respostas)

    acertos = 0

    for resposta in respostas:

        alternativa = db.query(Alternativa).filter(
            Alternativa.id == resposta.alternativa_id
        ).first()

        if alternativa and alternativa.correta:
            acertos += 1

    erros = total - acertos

    percentual = 0

    if total > 0:
        percentual = round((acertos / total) * 100, 2)

    resultado_existente = db.query(Resultado).filter(
        Resultado.tentativa_id == ultima_tentativa.id
    ).first()

    if not resultado_existente:

        novo_resultado = Resultado(
            usuario_id=usuario_id,
            total=total,
            acertos=acertos,
            erros=erros,
            percentual=percentual,
            tentativa_id=ultima_tentativa.id
        )

        db.add(novo_resultado)
        ultima_tentativa.concluida = True
        db.commit()

    return {
        "usuario_id": usuario_id,
        "tentativa_id": ultima_tentativa.id,
        "total": total,
        "acertos": acertos,
        "erros": erros,
        "percentual": percentual
    }
    
@app.get("/resultados")
def listar_resultados(
    db: Session = Depends(get_db)
):

    resultados = db.query(Resultado).all()

    return resultados

@app.get("/resultados/usuario/{usuario_id}")
def listar_resultados_usuario(
    usuario_id: int,
    db: Session = Depends(get_db)
):

    resultados = db.query(Resultado).filter(
        Resultado.usuario_id == usuario_id
    ).all()

    return resultados

@app.get("/quiz/{pergunta_id}")
def obter_quiz(
    pergunta_id: int,
    db: Session = Depends(get_db)
):

    pergunta = db.query(Pergunta).filter(
        Pergunta.id == pergunta_id
    ).first()

    if not pergunta:
        return JSONResponse(
            status_code=404,
            content={"erro": "Pergunta não encontrada"}
        )

    alternativas = db.query(Alternativa).filter(
        Alternativa.pergunta_id == pergunta_id
    ).all()

    random.shuffle(alternativas)
    lista_alternativas = []

    for alternativa in alternativas:

        lista_alternativas.append({
            "id": alternativa.id,
            "descricao": alternativa.descricao
        })

    return {
        "id": pergunta.id,
        "enunciado": pergunta.enunciado,
        "tipo": pergunta.tipo,
        "tema_id": pergunta.tema_id,
        "alternativas": lista_alternativas
    }

@app.post("/login")
def login(
    dados: LoginRequest,
    db: Session = Depends(get_db)
):

    usuario = db.query(Usuario).filter(
        Usuario.email == dados.email
    ).first()

    if not usuario:
        return {
            "mensagem": "Email ou senha inválidos"
        }

    if usuario.senha != dados.senha:
        return {
            "mensagem": "Email ou senha inválidos"
        }

    db.query(Resposta).filter(
        Resposta.usuario_id == usuario.id
    ).delete()

    db.commit()

    return {
        "mensagem": "Login realizado com sucesso",
        "usuario_id": usuario.id,
        "nome": usuario.nome,
        "admin": usuario.admin
    }

@app.get("/admin/resultados")
def admin_resultados(
    db: Session = Depends(get_db)
):

    resultados = db.query(Resultado).all()

    lista = []

    for resultado in resultados:

        usuario = db.query(Usuario).filter(
            Usuario.id == resultado.usuario_id
        ).first()

        print(dados.email)
        print (usuario)

        lista.append({
            "nome": usuario.nome,
            "acertos": resultado.acertos,
            "erros": resultado.erros,
            "percentual": resultado.percentual
        })

    return lista
    
@app.get("/ranking")
def ranking(
    db: Session = Depends(get_db)
):

    usuarios = db.query(Usuario).all()

    lista = []

    for usuario in usuarios:

        melhor_resultado = db.query(Resultado).filter(
            Resultado.usuario_id == usuario.id
        ).order_by(
            Resultado.percentual.desc()
        ).first()

        if melhor_resultado:

            lista.append({
                "nome": usuario.nome,
                "percentual": melhor_resultado.percentual
            })

    lista = sorted(
        lista,
        key=lambda x: x["percentual"],
        reverse=True
    )

    return lista


@app.get("/exportar-excel")
def exportar_excel(
    db: Session = Depends(get_db)
):

    workbook = Workbook()

    planilha = workbook.active

    planilha.title = "Resultados"

    planilha.append([
        "Aluno",
        "Acertos",
        "Erros",
        "Percentual"
    ])

    resultados = db.query(Resultado).all()

    for resultado in resultados:

        usuario = db.query(Usuario).filter(
            Usuario.id == resultado.usuario_id
        ).first()

        planilha.append([
            usuario.nome,
            resultado.acertos,
            resultado.erros,
            resultado.percentual
        ])

    arquivo = "resultado_alunos.xlsx"

    workbook.save(arquivo)

    return FileResponse(
    path=arquivo,
    filename=arquivo,
    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def normalizar_texto(texto):

    if texto is None:
        return ""

    return (
        str(texto)
        .lower()
        .replace('"', '')
        .replace("'", '')
        .strip()
    )
@app.get("/exportar-auditoria")
def exportar_auditoria(
    db: Session = Depends(get_db)
):

    workbook = Workbook()

    planilha = workbook.active

    planilha.title = "Auditoria"

    planilha.append([
        "Usuário",
        "Ação",
        "Data/Hora"
    ])

    logs = db.query(
        Auditoria
    ).order_by(
        Auditoria.data_hora.desc()
    ).all()

    for log in logs:

        usuario = db.query(
            Usuario
        ).filter(
            Usuario.id == log.usuario_id
        ).first()

        nome_usuario = "Sistema"

        if usuario:
            nome_usuario = usuario.nome

        planilha.append([
            nome_usuario,
            log.acao,
            log.data_hora.strftime(
                "%d/%m/%Y %H:%M:%S"
            )
        ])

    arquivo = "auditoria.xlsx"

    workbook.save(arquivo)

    return FileResponse(
        path=arquivo,
        filename=arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/importar-perguntas")
def importar_perguntas(
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db)
):

    workbook = load_workbook(
        arquivo.file
    )

    planilha = workbook.active

    importadas = 0

    ignoradas = 0

    for linha in planilha.iter_rows(
        min_row=2,
        values_only=True
    ):

        tema_nome = linha[0]
        enunciado = linha[1]

        alternativa_a = linha[2]
        alternativa_b = linha[3]
        alternativa_c = linha[4]
        alternativa_d = linha[5]
        alternativa_e = linha[6]

        correta = linha[7]

        if not enunciado:

            ignoradas += 1

            continue

        tema = db.query(Tema).filter(
            Tema.nome == tema_nome
        ).first()
        if not tema:

            tema = Tema(
                nome=tema_nome,
                liberado=False
            )

            db.add(tema)

            db.commit()

            db.refresh(tema)

        perguntas_tema = db.query(Pergunta).filter(
            Pergunta.tema_id == tema.id
        ).all()

        pergunta_existente = False

        enunciado_normalizado = normalizar_texto(
            enunciado
        )

        for pergunta in perguntas_tema:

            if (
                normalizar_texto(
                    pergunta.enunciado
                )
                ==
                enunciado_normalizado
            ):

                pergunta_existente = True

                break

        if pergunta_existente:

            ignoradas += 1

            continue

        nova_pergunta = Pergunta(
            enunciado=enunciado,
            tipo="objetiva",
            tema_id=tema.id
        )

        db.add(nova_pergunta)

        db.commit()

        db.refresh(nova_pergunta)

        alternativas = [
            ("A", alternativa_a),
            ("B", alternativa_b),
            ("C", alternativa_c),
            ("D", alternativa_d),
            ("E", alternativa_e)
        ]

        for letra, descricao in alternativas:

            nova_alternativa = Alternativa(
                descricao=descricao,
                correta=(letra == correta),
                pergunta_id=nova_pergunta.id
            )

            db.add(nova_alternativa)

        db.commit()

        importadas += 1

    novo_log = Auditoria(
        usuario_id=1,
        acao=f"Planilha importada: {importadas} perguntas",
        data_hora=datetime.now()
    )

    db.add(novo_log)

    db.commit()

    return {
        "mensagem":
        f"{importadas} perguntas importadas. {ignoradas} perguntas ignoradas."
    }

@app.get("/dashboard")
def dashboard(
    db: Session = Depends(get_db)
):

    total_usuarios = db.query(
        Usuario
    ).count()

    usuarios_ativos = db.query(
        Tentativa.usuario_id
    ).distinct().count()

    total_temas = db.query(
        Tema
    ).count()

    total_perguntas = db.query(
        Pergunta
    ).count()

    return {
        "usuarios": total_usuarios,
        "ativos": usuarios_ativos,
        "temas": total_temas,
        "perguntas": total_perguntas
    }
@app.get("/auditoria")
def listar_auditoria(
    db: Session = Depends(get_db)
):

    hoje = datetime.now().date()

    logs = db.query(
        Auditoria
    ).all()

    resultado = []

    for log in logs:

        if log.data_hora.date() != hoje:
            continue

        usuario = db.query(
            Usuario
        ).filter(
            Usuario.id == log.usuario_id
        ).first()

        nome_usuario = "Sistema"

        if usuario:
            nome_usuario = usuario.nome

        resultado.append({
            "id": log.id,
            "usuario": nome_usuario,
            "acao": log.acao,
            "data_hora": log.data_hora.strftime(
                "%d/%m/%Y %H:%M:%S"
            )
        })

    resultado.sort(
        key=lambda x: x["id"],
        reverse=True
    )

    return resultado

@app.get("/perguntas-tema-liberado")
def perguntas_tema_liberado(
    db: Session = Depends(get_db)
):

    tema = db.query(Tema).filter(
        Tema.liberado == True
    ).first()

    if not tema:

        return []

    perguntas = db.query(
        Pergunta
    ).filter(
        Pergunta.tema_id == tema.id
    ).all()

    return [
        {
            "id": pergunta.id,
            "enunciado": pergunta.enunciado
        }
        for pergunta in perguntas
    ]



@app.get("/quiz-tema")
def quiz_tema(
    db: Session = Depends(get_db)
):

    tema = db.query(Tema).filter(
        Tema.liberado == True
    ).first()

    if not tema:

        return {
            "erro":
            "Nenhum tema liberado."
        }

    perguntas = db.query(
        Pergunta
    ).filter(
        Pergunta.tema_id == tema.id
    ).all()

    return {
        "tema": tema.nome,
        "perguntas": len(perguntas)
    }
